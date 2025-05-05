import os
import can
import time
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook


from lutz_functions import *
from lutz_algos import *

from ascon import *



key_128 = b'5\x8d~\x96\xec^\xae\xca\x05\x0c\x02`*.\xf1o'  # Random 128-bit key (16 bytes)
key_168 = b'\xa7\x84$\xae\x86z}\x0cI\xe3kSy\xe1\xa8\xe3:*H\xc8E'  # Random 168-bit key (21 bytes)
key_192 = b"}\x94\x7f\xc7\xa8\xd39\x19\xc4,\xeb'\xf0\xf9\xba\xbb\xc6\xe0c\x97\xbbT\xa2/" # Random 192-bit key (24 btyes)
key_256 = b"0\xa1O\xe2\xf9K|\xeb\xe4\x18\xa2:\xb3\x92\xc0\x88dSq\xfb2\xc5G\xff'\x87\xec\x94\x11\x8a-\xf3" # Random 256-bit key (32 bytes)
key_320 = b'\xec$[\x94th.\xd3\x8b\x94\x03\x92\xc7\xbe\x14\r\xd8$i\xae\x82\xbb*\x9dc\x94pj\xeaS\xb2s\x81\x99bE\x0e\xfb\xdaN' # Random 320-bit key (40 bytes)
key_448 = b'F\x0e)S\x00\xf5\xf9\xba\xaf\xca\x8dgX7YLN\xd1\xa3q\x1dr;\xb9\xb1\t\xcf\xdc\xca \x88\xdbUT\xac\x12*\xf53|\xd4\x05\xc3^\xf6\xbb\xfb\x83\xe8%\xd7\x1e\x14\x92\x15f' # Random 448-bit key (56 bytes)

key_arb_pair_AES_CBC = [(key_128, 0xAEC128), (key_192, 0xAEC192), (key_256, 0xAEC256)] # Encryption key, arb_id pairs for AES CBC
key_arb_pair_AES_GCM = [(key_128, 0xAE128), (key_192, 0xAE192), (key_256, 0xAE256)] # Encryption key, arb_id pairs for AES GCM
key_arb_pair_Blowfish = [(key_128, 0xBF128), (key_256, 0xBF256), (key_320, 0xBF320), (key_448, 0xBF448)] # Encryption key, arb_id pairs for Blowfish CBC
key_arb_pair_Camellia = [(key_128, 0xCA128), (key_192, 0xCA192), (key_256, 0xCA256)] # Encryption key, arb_id pairs for Camellia
key_arb_pair_3DES = [(key_128, 0xDE128), (key_192, 0xDE192)] # Encryption key, arb_id pairs for 3DES

setup_physical_can_fd("can0")
bus = can.interface.Bus(channel='can0', interface='socketcan', fd=True) 
msg = can.Message(arbitration_id=0xD,  
                                    data= b"Start of Test" ,  # Data (max 8 bytes for standard CAN)
                                    is_fd=True)  # Set to True for extended CAN ID


bus.send(msg)
bus.shutdown()
set_down_physical_can("can0")

# filename_list = ["random_bytes_500KB.txt", "random_bytes_1MB.txt", "random_bytes_2MB.txt"]
filename_list = ["random_bytes_500KB.txt"]

# algo_list = ["Control", "AES_CBC", "AES_GCM", "Blowfish", "Camellia", "3DES"]
algo_list = ["AES_CBC", "AES_GCM", "Blowfish", "Camellia", "3DES"]

# Removing file used for encryption timing to start fresh
if os.path.exists("./Data/encryption_times.csv"):
    os.remove("./Data/encryption_times.csv")
    
# Removing file used for encryption timing to start fresh
if os.path.exists("./Data/cpu_load.csv"):
    os.remove("./Data/cpu_load.csv")
    

with open("./Data/cpu_load.csv", 'a') as cpu_load_writer:
    cpu_load_writer.write("Algorithm, Key Size, Load\n")
    
    with open("./Data/encryption_times.csv", 'a') as encryption_times_writer:
        encryption_times_writer.write("Algorithm, Key Size, Time\n")


        load_counter = 0
        for filename in filename_list:
            for algo in algo_list:

                byte_date_filename = filename
                byte_data_file_size = os.path.getsize(byte_date_filename)
                print(f"Testing with file of size: {byte_data_file_size} bytes")    
                
                print(f"Testing with bitrate of 1000000 bps")
                setup_physical_can_fd("can0")
                setup_physical_can_fd("can1")
                
                # Defines the CAN interface 
                bus = can.interface.Bus(channel='can0', interface='socketcan', fd=True) 

                
                
                if algo == "Control":
                    time_info_df = pd.DataFrame()
                    
                    totalRuns = 2
                    curRun = 1
                    curRunData = []
                    
                    print("Control test starting...")

                    while curRun <= totalRuns:
                        start = time.perf_counter()
                        
                        for can_message in read_file_in_chunks(byte_date_filename, 64): # Chunks data into 64 bytes
                            
                            # Create the CAN message
                            msg = can.Message(arbitration_id=0xC,  # CAN ID
                                            data= can_message ,  # Data (max 8 bytes for standard CAN)
                                            is_fd=True,         # Set to True for extended CAN ID
                                            is_extended_id=True)  # Allows extended arb id
                            
                            # Attempt to send message, if failed, retry
                            send_message(bus, msg)
                            
                            # Periodically measures the CPU load
                            if load_counter > 100:
                                log_python_process_cpu(cpu_load_writer,"0",algo)
                                load_counter = 0
                            else:
                                load_counter +=1
                            
                                
                        end = time.perf_counter()  - start
                        curRunData.append(end)
                        
                        # Sends can message with arb_id 0xF to signal time to file compare
                        msg = can.Message(arbitration_id=0xF,  # CAN ID
                                            data= can_message ,  # Data (max 8 bytes for standard CAN)
                                            is_fd=True,         # Set to True for extended CAN ID
                                            is_extended_id=True)  # Allows extended arb id
                        send_message(bus, msg)

                        print(f"Control run {curRun} took: {end} seconds")
                        curRun += 1

                
                    time_info_df["Control Data"] = curRunData
                            
                elif algo == "AES_CBC":
                    time_info_df = pd.DataFrame()   
                    
                    IV_len = 16 
                    print(f"Testing the {algo} algorithm")
                    
                    for key, arb_id in key_arb_pair_AES_CBC:
                        print(f"Testing arb_id: {hex(arb_id)}")
                        
                        totalRuns = 2
                        curRun = 1
                        curRunData = []

                        while curRun <= totalRuns:
                            start = time.perf_counter()
                            
                            for can_message in read_file_in_chunks(byte_date_filename, 64-1-IV_len): # Actual data chunks must be 64-1-IV_length so that exactly 64 bytes of data is encoded and sent
                                
                                
                                # Encrypt CAN message data
                                start_decrypt = time.perf_counter()
                                encrypted_can_message = encrypt_AES_CBC(can_message, key, iv_length=IV_len)
                                encryption_time_elapsed = time.perf_counter() - start_decrypt
                                encryption_times_writer.write(f"{algo},{str(len(key)*8)},{str(encryption_time_elapsed)}\n")
                                
                                # Create the CAN message
                                msg = can.Message(arbitration_id=arb_id,  # CAN ID
                                                data= encrypted_can_message ,  # Data (max 8 bytes for standard CAN)
                                                is_fd=True,         # Set to True for extended CAN ID
                                                is_extended_id=True)  # Allows extended arb id

                                # Attempt to send message, if failed, retry
                                send_message(bus, msg)
                                
                                # Periodically measures the CPU load
                                if load_counter > 100:
                                    log_python_process_cpu(cpu_load_writer,str(len(key)*8),algo)
                                    load_counter = 0
                                else:
                                    load_counter +=1
                                    
                            end = time.perf_counter()  - start
                            curRunData.append(end)
                            
                            compare_truth_2_sent_file_data(bus)

                            print(f"Run {curRun} took: {end} seconds")
                            curRun += 1

                    
                        time_info_df[str(hex(arb_id))] = curRunData
                                
                elif algo == "AES_GCM":
                    time_info_df = pd.DataFrame()   
                    
                    nonce_len = 12
                    tag_len = 16
                    print(f"Testing the {algo} algorithm")
                    
                    for key, arb_id in key_arb_pair_AES_GCM:
                        print(f"Testing arb_id: {hex(arb_id)}")
                        
                        totalRuns = 2
                        curRun = 1
                        curRunData = []

                        while curRun <= totalRuns:
                            start = time.perf_counter()
                            
                            for can_message in read_file_in_chunks(byte_date_filename, 64-nonce_len-tag_len): # Actual data chunks must be 64-1-IV_length so that exactly 64 bytes of data is encoded and sent
                                
                                # Encrypt CAN message data
                                start_decrypt = time.perf_counter()
                                encrypted_can_message = encrypt_AES_GCM(can_message, key)
                                encryption_time_elapsed = time.perf_counter() - start_decrypt
                                encryption_times_writer.write(f"{algo},{str(len(key)*8)},{str(encryption_time_elapsed)}\n")
                                
                                if len(can_message) < 36:
                                    ciphertext = encrypted_can_message
                                
                                # Create the CAN message
                                msg = can.Message(arbitration_id=arb_id,  # CAN ID
                                                data= encrypted_can_message ,  # Data (max 8 bytes for standard CAN)
                                                is_fd=True,         # Set to True for extended CAN ID
                                                is_extended_id=True)  # Allows extended arb id

                                # Attempt to send message, if failed, retry
                                send_message(bus, msg)
                                
                                # Periodically measures the CPU load
                                if load_counter > 100:
                                    log_python_process_cpu(cpu_load_writer,str(len(key)*8),algo)
                                    load_counter = 0
                                else:
                                    load_counter +=1
                                    
                            end = time.perf_counter()  - start
                            curRunData.append(end)
                            
                            compare_truth_2_sent_file_data(bus)

                            print(f"Run {curRun} took: {end} seconds")
                            curRun += 1

                    
                        time_info_df[str(hex(arb_id))] = curRunData
                        
                elif algo == "Blowfish":
                    time_info_df = pd.DataFrame()   
                    
                    print(f"Testing the {algo} algorithm")
                    IV_len = 8
                    BLOWFISH_BLOCK_SIZE = 8
                    
                    for key, arb_id in key_arb_pair_Blowfish:
                        print(f"Testing arb_id: {hex(arb_id)}")
                        
                        totalRuns = 2
                        curRun = 1
                        curRunData = []
                        while curRun <= totalRuns:
                            start = time.perf_counter()
                            
                            for can_message in read_file_in_chunks(byte_date_filename, 64 - IV_len - BLOWFISH_BLOCK_SIZE ): # Actual data chunks must be 64-1-IV_length so that exactly 64 bytes of data is encoded and sent
                                
                                # Encrypt CAN message data
                                start_decrypt = time.perf_counter()
                                encrypted_can_message = encrypt_blowfish(can_message, key, IV_len)
                                encryption_time_elapsed = time.perf_counter() - start_decrypt
                                encryption_times_writer.write(f"{algo},{str(len(key)*8)},{str(encryption_time_elapsed)}\n")
                                
                                # Create the CAN message
                                msg = can.Message(arbitration_id=arb_id,  # CAN ID
                                                data= encrypted_can_message ,  # Data (max 8 bytes for standard CAN)
                                                is_fd=True,         # Set to True for extended CAN ID
                                                is_extended_id=True)  # Allows extended arb id

                                # Attempt to send message, if failed, retry
                                send_message(bus, msg)
                                
                                # Periodically measures the CPU load
                                if load_counter > 100:
                                    log_python_process_cpu(cpu_load_writer,str(len(key)*8),algo)
                                    load_counter = 0
                                else:
                                    load_counter +=1
                                    
                            end = time.perf_counter()  - start
                            curRunData.append(end)
                            
                            compare_truth_2_sent_file_data(bus)

                            print(f"Run {curRun} took: {end} seconds")
                            curRun += 1

                    
                        time_info_df[str(hex(arb_id))] = curRunData
                            
                elif algo == "Camellia":   
                    time_info_df = pd.DataFrame() 
                    
                    IV_len = 16
                    print(f"Testing the {algo} algorithm")
                    
                    for key, arb_id in key_arb_pair_Camellia:
                        print(f"Testing arb_id: {hex(arb_id)}")
                        
                        totalRuns = 2
                        curRun = 1
                        curRunData = []

                        while curRun <= totalRuns:
                            start = time.perf_counter()
                            
                            for can_message in read_file_in_chunks(byte_date_filename, 64-1-IV_len): # Actual data chunks must be 64-1-IV_length so that exactly 64 bytes of data is encoded and sent
                                
                                
                                # Encrypt CAN message data
                                start_decrypt = time.perf_counter()
                                encrypted_can_message = encrypt_camellia(can_message, key, IV_len)
                                encryption_time_elapsed = time.perf_counter() - start_decrypt
                                encryption_times_writer.write(f"{algo},{str(len(key)*8)},{str(encryption_time_elapsed)}\n")
                                
                                # Create the CAN message
                                msg = can.Message(arbitration_id=arb_id,  # CAN ID
                                                data= encrypted_can_message ,  # Data (max 8 bytes for standard CAN)
                                                is_fd=True,         # Set to True for extended CAN ID
                                                is_extended_id=True)  # Allows extended arb id

                                # Attempt to send message, if failed, retry
                                send_message(bus, msg)
                                
                                # Periodically measures the CPU load
                                if load_counter > 100:
                                    log_python_process_cpu(cpu_load_writer,str(len(key)*8),algo)
                                    load_counter = 0
                                else:
                                    load_counter +=1
                                    
                            end = time.perf_counter()  - start
                            curRunData.append(end)
                            
                            compare_truth_2_sent_file_data(bus)

                            print(f"Run {curRun} took: {end} seconds")
                            curRun += 1

                    
                        time_info_df[str(hex(arb_id))] = curRunData
                
                elif algo == "3DES": 
                    time_info_df = pd.DataFrame()
                    
                    IV_len = 8   
                    print(f"Testing the {algo} algorithm")
                    for key, arb_id in key_arb_pair_3DES:
                        print(f"Testing arb_id: {hex(arb_id)}")
                        
                        totalRuns = 2
                        curRun = 1
                        curRunData = []

                        while curRun <= totalRuns:
                            start = time.perf_counter()
                            
                            for can_message in read_file_in_chunks(byte_date_filename, 48): # Actual data chunks must be 64-3-IV_length so that exactly 64 bytes of data is encoded and sent
                                
                                
                                # Encrypt CAN message data
                                start_decrypt = time.perf_counter()
                                encrypted_can_message = encrypt_3DES(can_message, key, IV_len)
                                encryption_time_elapsed = time.perf_counter() - start_decrypt
                                encryption_times_writer.write(f"{algo},{str(len(key)*8)},{str(encryption_time_elapsed)}\n")
                                
                                # Create the CAN message
                                msg = can.Message(arbitration_id=arb_id,  # CAN ID
                                                data= encrypted_can_message ,  # Data (max 8 bytes for standard CAN)
                                                is_fd=True,         # Set to True for extended CAN ID
                                                is_extended_id=True)  # Allows extended arb id

                                # Attempt to send message, if failed, retry
                                send_message(bus, msg)
                                
                                # Periodically measures the CPU load
                                if load_counter > 100:
                                    log_python_process_cpu(cpu_load_writer,str(len(key)*8),algo)
                                    load_counter = 0
                                else:
                                    load_counter +=1
                                
                            end = time.perf_counter()  - start
                            curRunData.append(end)
                            
                            compare_truth_2_sent_file_data(bus)

                            print(f"Run {curRun} took: {end} seconds")
                            curRun += 1

                    
                        time_info_df[str(hex(arb_id))] = curRunData
                        
                elif algo == "ChaCha20": 
                    time_info_df = pd.DataFrame()
                    
                    full_nonce_len = 16  
                    key = key_256
                    arb_id = 0xCC20
                    
                    print(f"Testing the {algo} algorithm")
                    
                    totalRuns = 2
                    curRun = 1
                    curRunData = []

                    while curRun <= totalRuns:
                        start = time.perf_counter()
                        
                        counter = 0
                        for can_message in read_file_in_chunks(byte_date_filename, 64-full_nonce_len): # Actual data chunks must be 64-1-IV_length so that exactly 64 bytes of data is encoded and sent
                            
                            
                            # Encrypt CAN message data
                            start_decrypt = time.perf_counter()
                            encrypted_can_message, counter = encrypt_chacha20(can_message, key, full_nonce_len, counter)
                            encryption_time_elapsed = time.perf_counter() - start_decrypt
                            encryption_times_writer.write(f"{algo},{str(len(key)*8)},{str(encryption_time_elapsed)}\n")
                            
                            # Create the CAN message
                            msg = can.Message(arbitration_id=arb_id,  # CAN ID
                                            data= encrypted_can_message ,  # Data (max 8 bytes for standard CAN)
                                            is_fd=True,         # Set to True for extended CAN ID
                                            is_extended_id=True)  # Allows extended arb id

                            # Attempt to send message, if failed, retry
                            send_message(bus, msg)
                            
                            # Periodically measures the CPU load
                            if load_counter > 100:
                                log_python_process_cpu(cpu_load_writer,str(len(key)*8),algo)
                                load_counter = 0
                            else:
                                load_counter +=1
                                
                        end = time.perf_counter()  - start
                        curRunData.append(end)
                        
                        compare_truth_2_sent_file_data(bus)

                        print(f"Run {curRun} took: {end} seconds")
                        curRun += 1
                    
                    time_info_df[algo] = curRunData
                
                elif algo == "Ascon": 
                    time_info_df = pd.DataFrame()
                    
                    nonce_len = 16  
                    tag_length = 16
                    key = key_128
                    arb_id = 0xAC128
                    
                    print(f"Testing the {algo} algorithm")
                    
                    totalRuns = 2
                    curRun = 1
                    curRunData = []

                    while curRun <= totalRuns:
                        start = time.perf_counter()

                        for can_message in read_file_in_chunks(byte_date_filename, 64-nonce_len-tag_length): # Actual data chunks must be 64-1-IV_length so that exactly 64 bytes of data is encoded and sent
                            
                            
                            # Encrypt CAN message data
                            start_decrypt = time.perf_counter()
                            encrypted_can_message = encrypt_ascon128(can_message, key, associated_data = b'')
                            encryption_time_elapsed = time.perf_counter() - start_decrypt
                            encryption_times_writer.write(f"{algo},{str(len(key)*8)},{str(encryption_time_elapsed)}\n")
                            
                            # Create the CAN message
                            msg = can.Message(arbitration_id=arb_id,  # CAN ID
                                            data= encrypted_can_message ,  # Data (max 8 bytes for standard CAN)
                                            is_fd=True,         # Set to True for extended CAN ID
                                            is_extended_id=True)  # Allows extended arb id

                            # Attempt to send message, if failed, retry
                            send_message(bus, msg)
                            
                            # Periodically measures the CPU load
                            if load_counter > 100:
                                log_python_process_cpu(cpu_load_writer,str(len(key)*8),algo)
                                load_counter = 0
                            else:
                                load_counter +=1
                                
                        end = time.perf_counter()  - start
                        curRunData.append(end)
                        
                        compare_truth_2_sent_file_data(bus)

                        print(f"Run {curRun} took: {end} seconds")
                        curRun += 1
                    
                    time_info_df[algo] = curRunData
                
                if os.path.exists(f'./Data/{algo}_{byte_data_file_size}-bytes.xlsx'):
                    with pd.ExcelWriter(f'./Data/{algo}_{byte_data_file_size}-bytes.xlsx', engine='openpyxl', mode='a') as writer:
                        time_info_df.to_excel(writer, sheet_name=f'{algo}_DATA', index=False)
                    
                else:
                    time_info_df.to_excel(f'./Data/{algo}_{byte_data_file_size}-bytes.xlsx',sheet_name=f'{algo}_DATA', index=False)

                # Load Excel workbook made earlier with dataframe to add specific data cells
                wb = load_workbook(f'./Data/{algo}_{byte_data_file_size}-bytes.xlsx')
                
                
                
                # Iterates through each row in the dataframe to give stats on each one as
                # well as creates a seperate stat sheet for each key length
                for column_name, column_data in time_info_df.items():
                    stats_sheet = wb.create_sheet(title=f'{str(column_name)}_Stats')
                    
                    stats_sheet['A1'] = "Average"
                    stats_sheet['A2'] = np.mean(column_data)
                    
                    stats_sheet['B1'] = "Minimum Value"
                    stats_sheet['B2'] = minVal = np.min(column_data)
                    
                    stats_sheet['C1'] = "Maximum Value"
                    stats_sheet['C2'] = maxVal = np.max(column_data)
                

                # Save the workbook as an .xlsx file
                wb.save(f'./Data/{algo}_{byte_data_file_size}-bytes.xlsx')
                wb.close()
                
                if algo == algo_list[-1] and filename == filename_list[-1]:
                    msg = can.Message(arbitration_id=0xD,  # 0xD arb id used to tell receive side to finish testing
                                                data= os.urandom(64) ,  # Data (max 8 bytes for standard CAN)
                                                is_fd=True)  # Set to True for extended CAN ID
                    bus.send(msg,1)
                    
                else:
                    msg = can.Message(arbitration_id=0xA,  # 0xA arb id used to tell receive side to restart receiving process for next algorithm
                                                data= os.urandom(64) ,  # Data (max 8 bytes for standard CAN)
                                                is_fd=True)  # Set to True for extended CAN ID
                    print(f"sending final message of this test")
                    
                    # Continually send a message to indicate the Rx side to start over
                    while True:
                        result = subprocess.run(
                            ["ip", "link", "show", "can1"],
                            capture_output=True,
                            text=True,
                            check=True
                        )
                        if "state UP" in result.stdout:
                            bus.send(msg,1)
                            continue
                        else:
                            break
                
                
                bus.shutdown()
                set_down_physical_can("can0")
                print("\n")
                
        






